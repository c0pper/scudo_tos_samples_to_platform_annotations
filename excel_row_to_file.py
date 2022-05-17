from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

category = "a"

wb_name = "clauses_types"
wb = load_workbook(f'{wb_name}.xlsx')

ws = wb["A"]
clause_col = ws['C']

def create_clauses_dict():
    clauses = []
    print(get_column_letter(1))

    for cell in clause_col:
        reasons = []
        reason_startingcol = 4
        while ws[f'{get_column_letter(reason_startingcol)}{cell.row}'].value:
            reason_cell = f'{get_column_letter(reason_startingcol)}{cell.row}'
            reasons.append(ws[reason_cell].value)
            reason_startingcol += 1

        if ws[f'B{cell.row}'].value[-1] == "2":
            grade = f"{category}2"
            cat_grade_reason = {"cat": category, "grade": grade, "reasons": reasons, "text": cell.value}
            clauses.append(cat_grade_reason)
        elif ws[f'B{cell.row}'].value[-1] == "3":
            grade = f"{category}3"
            cat_grade_reason = {"cat": category, "grade": grade, "reasons": reasons, "text": cell.value}
            clauses.append(cat_grade_reason)
        else:
            grade = f"{category}1"
            cat_grade_reason = {"cat": category, "grade": grade, "reasons": reasons, "text": cell.value}
            clauses.append(cat_grade_reason)
    
    return clauses

clauses = create_clauses_dict()
# print(clauses)

tpath = f"single_clauses_{category}/test/"
apath = f"single_clauses_{category}/ann/"
for idx, clause in enumerate(clauses):
    # print(path + f'{idx}.txt')
    with open(tpath + f'{idx}.txt', 'w', encoding="UTF8") as f:
        f.write(clause["text"])
    with open(apath + f'{idx}.ann', 'w', encoding="UTF8") as f:
        for idx, r in enumerate(clause["reasons"]):
            f.write(f"C{idx+1}		{r}\n")
        print(clause['grade'])
        f.write(f"C{idx+1}		{clause['grade']}\n")
