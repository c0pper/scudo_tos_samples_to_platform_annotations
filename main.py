from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

category = "Unilateral termination"

wb_name = "unilateral_term"
wb = load_workbook(f'{wb_name}.xlsx')

ws = wb.active
clause_col = ws['C']

def create_clauses_dict():
    clauses = []
    print(get_column_letter(1))

    for cell in clause_col:
        reasons = []
        reason_startingcol = 4
        while ws[f'{get_column_letter(reason_startingcol)}{cell.row}'].value:
            reason_cell = f'{get_column_letter(reason_startingcol)}{cell.row}'
            reasons.append(ws[reason_cell].value)
            reason_startingcol += 1

        if ws[f'B{cell.row}'].value[-1] == "2":
            grade = "Potentially Unfair"
            cat_grade_reason = {"cat": category, "grade": grade, "reasons": reasons, "text": cell.value}
            clauses.append(cat_grade_reason)
        elif ws[f'B{cell.row}'].value[-1] == "3":
            grade = "Unfair"
            cat_grade_reason = {"cat": category, "grade": grade, "reasons": reasons, "text": cell.value}
            clauses.append(cat_grade_reason)
        else:
            grade = "Fair"
            cat_grade_reason = {"cat": category, "grade": grade, "reasons": reasons, "text": cell.value}
            clauses.append(cat_grade_reason)
    
    return clauses

clauses = create_clauses_dict()
new = Workbook()
ws = new.active

r = 1
id_count = 0
for clause in clauses:
    # print(len(clause["reasons"]), clause["reasons"], clause["text"])

    n_reasons = len(clause["reasons"])

    # template id
    ws.cell(row=r, column=1).value = id_count
    ws.cell(row=r+1, column=1).value = id_count
    for n in range(n_reasons):
        ws.cell(row=r+2+n, column=1).value = id_count

    # fields
    ws.cell(row=r, column=2).value = "category"
    ws.cell(row=r+1, column=2).value = "grade"
    for n in range(n_reasons):
        ws.cell(row=r+2+n, column=2).value = "reason"

    #cpk values
    ws.cell(row=r, column=3).value = clause["cat"]
    ws.cell(row=r+1, column=3).value = clause["grade"]
    for n in range(n_reasons):
        ws.cell(row=r+2+n, column=3).value = clause["reasons"][n]

    r += 2+n_reasons
    id_count += 1

new.save(f'{wb_name}_targets.xlsx')